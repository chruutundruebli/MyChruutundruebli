from openpyxl import Workbook

# import base64
# import hmac
# import hashlib
# from urllib import parse
# from django.conf import settings

# from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse#, HttpResponseRedirect, JsonResponse

# from juntagrico.models import Member, Subscription

# from juntagrico.dao.depotdao import DepotDao
# from juntagrico.dao.listmessagedao import ListMessageDao
# from juntagrico.dao.subscriptionsizedao import SubscriptionSizeDao
# from juntagrico.util.pdf import render_to_pdf_http
# from juntagrico.util.temporal import weekdays, start_of_business_year, end_of_business_year

import juntagrico.dao
from juntagrico.config import Config
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from openpyxl.utils import get_column_letter


# API

def excel_export_subscriptions(request):
    filename = '{}_{}.xlsx'.format(Config.vocabulary('subscription_pl'), timezone.now().date())
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = Workbook()

    # Sheet 1: Subscriptions with prices
    ws1 = wb.active
    ws1.title = Config.vocabulary('subscription_pl')

    # header
    ws1.cell(1, 1, u"{}".format(Config.vocabulary('member_pl')))
    ws1.column_dimensions['A'].width = 40
    ws1.cell(1, 2, u"{}".format(_('E-Mail')))
    ws1.column_dimensions['B'].width = 30
    ws1.cell(1, 3, u"{}".format(_('Gesamtpreis [{}]').format(Config.currency())))
    ws1.column_dimensions['C'].width = 17
    for column, subs_type in enumerate(SubscriptionTypeDao.get_all(), 4):
        ws1.cell(1, column, u"EAT {}".format(subs_type.price))
        ws1.column_dimensions[get_column_letter(column)].width = 17

    # data
    for row, subscription in enumerate(SubscriptionDao.all_active_subscritions(), 2):
        ws1.cell(row, 1, ", ".join([member.get_name() for member in subscription.members.all()]))
        ws1.cell(row, 2, subscription.primary_member.email)
        ws1.cell(row, 3, subscription.price)
        for column, subs_type in enumerate(SubscriptionTypeDao.get_all(), 4):
            ws1.cell(row, column, subscription.types.filter(id=subs_type.id).count())

    ws1.freeze_panes = ws1['A2']
    wb.save(response)
    return response
